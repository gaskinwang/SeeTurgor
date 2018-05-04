import numpy as np
import scipy.io

Path = scipy.io.loadmat('prepare.mat')['Path'].astype(int)
ColorAllocation = np.squeeze(scipy.io.loadmat('ColorAllocation.mat')
                             ['ColorAllocation'].astype(int))

Boundary = np.arange(1, 14)
Levals = []
Levals.append(Boundary.tolist())  # the 0 layer
known = Boundary.tolist()
# Waiting for optimization
sumtemp = []
sumcolor = []
for la in range(1, 7):
    temp = []
    # Waiting for optimization
    color_temp = []
    for cellnum in np.arange(Path.max()) + 1:
        if cellnum in known:
            pass
        else:
            for pat in Path:
                pat_ = pat.tolist()
                if (cellnum in pat_) and (pat_[pat_.index(cellnum) - 1] in known):
                    temp.append(cellnum)
                    # Waiting for optimization
                    color_temp.append(ColorAllocation[cellnum-1])
                    break

    known.extend(temp)
    Levals.append(temp)
    print('debug, Leval', la, ':', temp)
    print('debug, Color', la, ':', color_temp)
    
    
    sumtemp.append(temp)
    sumcolor.append(color_temp)

remove_list = np.ones(Path.shape[0]) == 0
for path_num, pat in enumerate(Path):
    for ind, pat_cell_num in enumerate(pat):
        if pat_cell_num in Levals[ind]:
            pass
        else:
            remove_list[path_num] = True
            break
remained_path = Path[~ remove_list]
print('all the correct path: \n', remained_path)
scipy.io.savemat('Leval.mat', {('l' + str(ind)): np.array(lav) for ind, lav in enumerate(Levals)})
res = ColorAllocation[remained_path - 1]
res = res.astype(float)
res[res == 0] = 0.5

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
for ind1 in range(res.shape[0]):
    for ind2 in range(res.shape[1]):
        ws.cell(row=ind1 + 1, column=ind2 + 1, value=res[ind1, ind2])
wb.save('all the correct path4.xlsx')

wb = Workbook()
ws = wb.active
# for ind1 in range(len(sumtemp)):
#     for ind2 in range(len(sumtemp[0])):
#         ws.cell(row=ind1 + 1, column=ind2 + 1, value=sumtemp[ind1][ind2])
for ind1 in range(len(sumcolor)):
    for ind2 in range(len(sumcolor[ind1])):
        ws.cell(row=ind1 + 1, column=ind2 + 1, value=sumcolor[ind1][ind2])
wb.save('values and colors4.xlsx')
